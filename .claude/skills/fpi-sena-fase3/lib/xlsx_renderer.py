"""
xlsx_renderer.py — Renderer mecánico de plantilla GFPI-F-134_Vf.xlsx desde pm-3-7.json.

Per PLAN-FASE-3-ARQUITECTURA.md v1.4 §7 Hito-Fase3-4 task #5 + PM-3.7 master prompt v1.0
§XLSX RENDERER REQUIREMENTS.

Naturaleza: MECÁNICO determinístico · NO LLM · NO síntesis. Consume:
- `master-prompts/canon/GFPI-F-134_Vf.xlsx` (template oficial SENA April 2022 · read-only)
- `runs/<RUN-ID>/pm-3-7.json` (datos organizados 14 cols + metadata · producido por LLM via
  subagente_pm_3_7_gfpi_f134_matrix.py)

Produce: `runs/<RUN-ID>/pm-3-7-gfpi-f134-matrix.xlsx` con hoja 2 "PLANEACIÓN" poblada celdas
metadata (R4-R21) + data rows (R25+) preservando merged cells + styles del template.

Separation of concerns canónica (PM-3.7 master prompt v1.0):
- LLM (wrapper subagente_pm_3_7) = creative aggregation + transformación pedagógica → JSON
- xlsx renderer (este módulo) = mechanical template population → xlsx file
- NO mezclar responsabilidades · LLM no toca xlsx · renderer no inventa contenido

Cell coordinates canónicas (template GFPI-F-134_Vf.xlsx hoja 2):

Metadata block (rows 4-21):
  E4  = Fecha de Elaboración          (merged E4:N4)
  E5  = Programa nombre               (merged E5:N5)
  E6  = Modalidad                     (merged E6:N6)
  E7  = Código + versión              (merged E7:N7)
  E8  = Nombre Proyecto               (merged E8:N8)
  E9  = Código Proyecto               (merged E9:N9)
  E10:I10 + J10:N10 = Instructor 1 (Nombres + Regional)
  E11:I11 + J11:N11 = Instructor 2
  ... hasta E21:I21 + J21:N21 = Instructor 12 (max 12 instructores · ws expande si más)

Headers (rows 22-24): NO TOCAR · preservados del template

Data rows (rows 25+):
  Col A  = col_1_fase_proyecto                   (merged A25:A66 · single value cubre todo data)
  Col B  = col_2_actividad_proyecto              (merged B25:B47)
  Col C  = col_3_competencia                     (merged C25:C30 default · expandible)
  Col D  = col_4_resultado_aprendizaje
  Col E  = col_5_actividades_aprendizaje_a_desarrollar
  Col F  = col_6_horas_trabajo_directo
  Col G  = col_7_horas_trabajo_independiente
  Col H  = col_8_estrategias_didacticas_activas
  Col I  = col_9_ambiente                        (merged I25:K30 default vía AMBIENTES TIPIFICADOS group)
  Col J  = col_10_materiales_formacion
  Col K  = col_11_instructores_responsables
  Col L  = col_12_criterios_evaluacion
  Col M  = col_13_descripcion_evidencia_aprendizaje
  Col N  = col_14_observaciones

v1.0 SCOPE: single-RAP single-row scenario (single-guía absorpción common case · 1 RAP per
guide). Multi-RAP multi-row expansion = v1.1+ futuro.
"""

import json
from copy import copy
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    raise ImportError(
        "openpyxl required for xlsx_renderer. Install: pip install openpyxl --break-system-packages"
    ) from e


# ═══════════════════════════════════════════════════════════════════════════
# Cell coordinates canónicas (template GFPI-F-134_Vf.xlsx hoja 2)
# ═══════════════════════════════════════════════════════════════════════════

METADATA_CELLS = {
    "fecha_elaboracion": "E4",
    "programa_nombre": "E5",
    "modalidad_ejecucion": "E6",
    "codigo_version_programa": "E7",
    "nombre_proyecto": "E8",
    "codigo_proyecto": "E9",
}

# Instructor block: rows 10-21 (12 max) · cols E (nombre) + J (regional)
INSTRUCTOR_NOMBRE_COL = "E"
INSTRUCTOR_REGIONAL_COL = "J"
INSTRUCTOR_FIRST_ROW = 10
INSTRUCTOR_MAX = 12  # rows 10-21

# Data rows start at 25
DATA_ROW_START = 25

# 14 cols xlsx Vf hoja 2 → letras col
DATA_COLUMN_LETTERS = {
    "col_1_fase_proyecto": "A",
    "col_2_actividad_proyecto": "B",
    "col_3_competencia": "C",
    "col_4_resultado_aprendizaje": "D",
    "col_5_actividades_aprendizaje_a_desarrollar": "E",
    "col_6_horas_trabajo_directo": "F",
    "col_7_horas_trabajo_independiente": "G",
    "col_8_estrategias_didacticas_activas": "H",
    "col_9_ambiente": "I",
    "col_10_materiales_formacion": "J",
    "col_11_instructores_responsables": "K",
    "col_12_criterios_evaluacion": "L",
    "col_13_descripcion_evidencia_aprendizaje": "M",
    "col_14_observaciones": "N",
}

# Default canon paths
CANON_TEMPLATE_PATH = "master-prompts/canon/GFPI-F-134_Vf.xlsx"


# ═══════════════════════════════════════════════════════════════════════════
# Public API
# ═══════════════════════════════════════════════════════════════════════════

def render_gfpi_f134_matrix(pm37_data, pm0_context, output_path,
                              template_path=CANON_TEMPLATE_PATH, repo_root=None):
    """Render GFPI-F-134_Vf.xlsx hoja 2 from pm-3-7.json data + pm-0-context metadata.

    Args:
        pm37_data: dict · pm-3-7.json content (14 cols data + metadata)
        pm0_context: dict · pm-0-context.json content (programa metadata)
        output_path: str · path donde se guarda el xlsx output (NO sobrescribir template)
        template_path: str · default master-prompts/canon/GFPI-F-134_Vf.xlsx
        repo_root: str opcional · si template_path es relativo, prepend este

    Returns:
        dict {
            output_path: str,
            template_used: str,
            metadata_cells_written: int,
            data_cells_written: int,
            instructor_rows_written: int,
            validation_post_render: dict,
        }

    Raises:
        FileNotFoundError si template no existe
        ValueError si pm37_data missing critical fields
    """
    # Resolve paths
    template = Path(template_path)
    if repo_root and not template.is_absolute():
        template = Path(repo_root) / template
    if not template.exists():
        raise FileNotFoundError(f"Template canon no encontrado: {template}")

    output = Path(output_path)
    if repo_root and not output.is_absolute():
        output = Path(repo_root) / output

    # Anti-overwrite check (REGLA 7 master prompt PM-3.7)
    if output.resolve() == template.resolve():
        raise ValueError(
            f"REGLA 7 violation: output path == template path · NUNCA sobrescribir canon. "
            f"output={output} · template={template}"
        )

    # Load template
    wb = openpyxl.load_workbook(str(template), data_only=False)
    if "2. PLANEACIÓN" not in wb.sheetnames:
        raise ValueError(
            f"Template {template} NO tiene hoja '2. PLANEACIÓN' · sheets disponibles: {wb.sheetnames}"
        )
    ws = wb["2. PLANEACIÓN"]

    # 1. Populate metadata block (rows 4-9)
    metadata_written = _populate_metadata(ws, pm37_data, pm0_context)

    # 2. Populate instructor block (rows 10-21 max 12)
    instructors_written = _populate_instructors(ws, pm0_context, pm37_data)

    # 3. Populate data row(s) starting row 25
    data_written = _populate_data_rows(ws, pm37_data)

    # 4. Save (NEVER overwrite template)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))

    # 5. Post-render validation: re-read + verify
    validation = _validate_post_render(str(output), pm37_data)

    return {
        "output_path": str(output),
        "template_used": str(template),
        "metadata_cells_written": metadata_written,
        "data_cells_written": data_written,
        "instructor_rows_written": instructors_written,
        "validation_post_render": validation,
    }


# ═══════════════════════════════════════════════════════════════════════════
# Helpers privados
# ═══════════════════════════════════════════════════════════════════════════

def _populate_metadata(ws, pm37_data, pm0_context):
    """Populate metadata cells E4-E9. Returns count of cells written."""
    # Source mapping: prefer pm37_data fields · fallback pm0_context
    metadata_values = {
        "fecha_elaboracion": pm37_data.get("generated_date", ""),
        "programa_nombre": pm0_context.get("programa_nombre", ""),
        "modalidad_ejecucion": pm0_context.get("modalidad", "Presencial"),  # default
        "codigo_version_programa": _format_codigo_version(pm0_context),
        "nombre_proyecto": pm37_data.get("col_2_actividad_proyecto", "") or pm0_context.get("nombre_proyecto", ""),
        "codigo_proyecto": pm0_context.get("codigo_proyecto", "") or "",
    }

    written = 0
    for key, cell_coord in METADATA_CELLS.items():
        value = metadata_values.get(key, "")
        if value:
            ws[cell_coord] = str(value)
            written += 1
    return written


def _format_codigo_version(pm0_context):
    """Format código + versión per SENA convention: '<codigo> versión <X>'."""
    codigo = pm0_context.get("programa_codigo_sofia", "")
    # Try to extract version from various sources
    version = pm0_context.get("programa_version", "")
    if not version:
        version = pm0_context.get("fase_0_metadata", {}).get("programa_version", "")
    if not version:
        version = "1"  # default
    return f"{codigo} versión {version}".strip()


def _populate_instructors(ws, pm0_context, pm37_data):
    """Populate instructor block rows 10-21. Returns rows written count."""
    # Try pm37 data first · fallback pm0
    instructors = pm37_data.get("col_11_instructores_responsables")
    if isinstance(instructors, str):
        # Single instructor as string · convert to list
        instructors = [{"nombre": instructors, "regional": ""}]
    elif isinstance(instructors, list) and instructors and isinstance(instructors[0], str):
        # List of strings · convert
        instructors = [{"nombre": name, "regional": ""} for name in instructors]
    elif not isinstance(instructors, list):
        instructors = []

    # Fallback to pm0_context.instructor metadata
    if not instructors:
        instructor_field = pm0_context.get("instructor", "")
        if instructor_field:
            instructors = [{"nombre": str(instructor_field), "regional": ""}]

    rows_written = 0
    for idx, instr in enumerate(instructors[:INSTRUCTOR_MAX]):
        row = INSTRUCTOR_FIRST_ROW + idx
        if isinstance(instr, dict):
            nombre = instr.get("nombre", "") or instr.get("name", "")
            regional = instr.get("regional", "") or instr.get("centro_formacion", "")
        else:
            nombre = str(instr)
            regional = ""

        if nombre:
            # Write to merged cell top-left (E10, E11, ...)
            nombre_cell = f"{INSTRUCTOR_NOMBRE_COL}{row}"
            ws[nombre_cell] = f"Nombres y Apellidos\n\n{nombre}"
            rows_written += 1

            if regional:
                regional_cell = f"{INSTRUCTOR_REGIONAL_COL}{row}"
                ws[regional_cell] = f"Regional y Centro de formación\n\n{regional}"

    return rows_written


def _populate_data_rows(ws, pm37_data):
    """Populate data row at DATA_ROW_START with 14 cols. Returns cells written count.

    v1.0: single-row · single-RAP scenario.
    """
    written = 0
    for field_name, col_letter in DATA_COLUMN_LETTERS.items():
        value = pm37_data.get(field_name)
        if value is None or value == "":
            continue

        # Format value · str/int/list
        if isinstance(value, list):
            # Join list with newlines for readability in cell
            formatted = "\n".join(str(v) for v in value if v)
        elif isinstance(value, dict):
            # Nested dict · serialize as readable lines (key: value pairs)
            formatted = "\n".join(f"{k}: {v}" for k, v in value.items() if v)
        else:
            formatted = str(value)

        cell_coord = f"{col_letter}{DATA_ROW_START}"
        ws[cell_coord] = formatted
        written += 1

    return written


def _validate_post_render(output_path, pm37_data):
    """Re-read output xlsx + verify critical cells written. Returns validation dict."""
    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws = wb["2. PLANEACIÓN"]

    checks = {}

    # Check 1: critical data cells populated
    critical_cols = ["col_3_competencia", "col_4_resultado_aprendizaje",
                     "col_6_horas_trabajo_directo", "col_7_horas_trabajo_independiente",
                     "col_12_criterios_evaluacion", "col_13_descripcion_evidencia_aprendizaje"]
    populated = 0
    for col_field in critical_cols:
        col_letter = DATA_COLUMN_LETTERS[col_field]
        cell_value = ws[f"{col_letter}{DATA_ROW_START}"].value
        if cell_value:
            populated += 1
    checks["critical_cells_populated"] = f"{populated}/{len(critical_cols)}"
    checks["critical_cells_pass"] = populated == len(critical_cols)

    # Check 2: metadata block has fecha + programa
    checks["fecha_present"] = bool(ws["E4"].value)
    checks["programa_present"] = bool(ws["E5"].value)

    # Check 3: header row 23 untouched (canonical headers preserved)
    header_a23 = ws["A23"].value
    checks["header_a23_preserved"] = header_a23 and "FASE DE PROYECTO" in str(header_a23)

    # Check 4: file readable + sheet present
    checks["sheet_2_planeacion_present"] = "2. PLANEACIÓN" in wb.sheetnames

    # Aggregate veredicto
    checks["veredicto"] = "PASS" if all([
        checks["critical_cells_pass"],
        checks["fecha_present"],
        checks["programa_present"],
        checks["header_a23_preserved"],
        checks["sheet_2_planeacion_present"],
    ]) else "FAIL"

    return checks




# ═══════════════════════════════════════════════════════════════════════════
# V04 CANON · Cell coordinates (template GFPI-F-134_V04.xlsx · 1 hoja PLANEACIÓN POR RAPS)
# ═══════════════════════════════════════════════════════════════════════════
# Canon V04 oficial SENA · 2026-04-30 · multi-RAP shape (master prompt PM-3.7 v2.0 REGLA 12)

CANON_V04_TEMPLATE_PATH = "master-prompts/canon/GFPI-F-134_V04.xlsx"

# Metadata cells V04 (from V04 IMARPOR ground truth · Sergio fill-in 2026-04-30)
V04_METADATA_CELLS = {
    "fecha_elaboracion": "E6",
    "denominacion": "E7",        # Denominación del Programa
    "modalidad": "E8",            # Modalidad de Formación
    "codigo_version": "E9",       # Código + versión Programa
    "instructor_1": "E10",        # Primer integrante diseño
    "instructor_2": "E11",        # Segundo integrante diseño
}

# RAP → row mapping V04 (Sergio fill convention · 1 RAP por row no consecutivos)
V04_RAP_ROW = {
    "RA1": 15,
    "RA2": 18,
    "RA3": 20,
    "RA4": 21,
}

# 14 cols xlsx V04 → indices (R13-R14 headers · R15+ data)
V04_DATA_COLUMN_INDICES = {
    "col_1_competencia": 1,                          # C1: COMPETENCIA
    "col_2_resultado_aprendizaje": 2,                # C2: RAP
    "col_3_saberes_conceptos_y_principios": 4,       # C4: SABERES CONCEPTOS Y PRINCIPIOS
    "col_4_saberes_procesos": 5,                     # C5: SABERES DE PROCESO
    "col_5_criterios_evaluacion": 6,                 # C6: CRITERIOS EVALUACIÓN
    "col_6_actividades_aprendizaje_a_desarrollar": 7,# C7: ACTIVIDADES APRENDIZAJE
    "col_7_horas_trabajo_directo": 8,                # C8: HORAS TRABAJO DIRECTO
    "col_8_horas_trabajo_independiente": 9,          # C9: HORAS TRABAJO INDEPENDIENTE
    "col_9_descripcion_evidencia_aprendizaje": 10,   # C10: DESCRIPCIÓN EVIDENCIA
    "col_10_estrategias_didacticas_activas": 11,     # C11: ESTRATEGIAS DIDÁCTICAS
    "col_11_ambiente": 12,                           # C12: AMBIENTE
    "col_12_materiales_formacion": 13,               # C13: MATERIALES FORMACIÓN
    "col_13_instructores_responsables": 14,          # C14: INSTRUCTORES RESPONSABLES
    "col_14_observaciones": 15,                      # C15: OBSERVACIONES
}


# ═══════════════════════════════════════════════════════════════════════════
# V04 RENDERER · Public API
# ═══════════════════════════════════════════════════════════════════════════

def render_gfpi_f134_v04_matrix(pm37_data, output_path,
                                  template_path=CANON_V04_TEMPLATE_PATH, repo_root=None):
    """Render GFPI-F-134 V04 xlsx (1 hoja PLANEACIÓN POR RAPS · multi-RAP) from pm-3-7.json v2.0.

    Canon V04 oficial SENA (master prompt PM-3.7 v2.0 REGLAS 8-15) · 2026-04-30.
    Differs from render_gfpi_f134_matrix (Vf legacy · single-row) en:
    - Template: GFPI-F-134_V04.xlsx (1 hoja vs Vf 5 hojas)
    - Shape: multi-RAP rows[] (vs Vf single-row)
    - Cells: R15/R18/R20/R21 (vs Vf R25)
    - Metadata: E6-E11 (vs Vf E4-E9)
    - SUM formulas R22 preservadas

    Args:
        pm37_data: dict · pm-3-7.json v2.0 content (rows[] + programa_metadata + totals)
        output_path: str · path donde se guarda el xlsx output (NO sobrescribir template)
        template_path: str · default master-prompts/canon/GFPI-F-134_V04.xlsx
        repo_root: str opcional · si template_path es relativo, prepend este

    Returns:
        dict {
            output_path, template_used, schema_version_consumed, format_canon_consumed,
            metadata_cells_written, rap_rows_written, data_cells_written,
            sum_formulas_preserved, merged_ranges_preserved,
            validation_post_render
        }

    Raises:
        FileNotFoundError si template no existe
        ValueError si pm37_data shape != v2.0 (rows[] missing) o output == template
    """
    import shutil

    # Validate shape v2.0
    schema_version = pm37_data.get("schema_version", "")
    format_canon = pm37_data.get("format_canon", "")
    rows = pm37_data.get("rows", [])

    if schema_version != "v2.0":
        raise ValueError(
            f"render_gfpi_f134_v04_matrix requires pm-3-7.json schema_version='v2.0'. "
            f"Got: '{schema_version}'. Use render_gfpi_f134_matrix for v1.0/Vf legacy."
        )
    if format_canon != "V04":
        raise ValueError(
            f"render_gfpi_f134_v04_matrix requires format_canon='V04'. "
            f"Got: '{format_canon}'."
        )
    if not isinstance(rows, list) or not rows:
        raise ValueError("pm-3-7.json rows[] missing or empty (REGLA 9 · multi-RAP shape required)")

    # Resolve paths
    template = Path(template_path)
    if repo_root and not template.is_absolute():
        template = Path(repo_root) / template
    if not template.exists():
        raise FileNotFoundError(f"V04 template canon no encontrado: {template}")

    output = Path(output_path)
    if repo_root and not output.is_absolute():
        output = Path(repo_root) / output

    # Anti-overwrite (REGLA 7 master prompt)
    if output.resolve() == template.resolve():
        raise ValueError(
            f"REGLA 7 violation: output path == template path · NUNCA sobrescribir canon V04. "
            f"output={output} · template={template}"
        )

    # Copy empty template (preserves merged cells + styles + SUM formulas)
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(str(template), str(output))

    # Open output and populate
    wb = openpyxl.load_workbook(str(output), data_only=False)
    ws = wb.active  # PLANEACIÓN POR RAPS (single sheet)

    # 1. Populate metadata cells (E6-E11)
    metadata_written = _populate_v04_metadata(ws, pm37_data)

    # 2. Populate data rows (one per RAP per V04_RAP_ROW mapping)
    rap_rows_written, data_cells_written = _populate_v04_data_rows(ws, rows)

    # 3. Verify SUM formulas R22 preserved (sanity check · we don't write to R22)
    sum_h22 = ws.cell(22, 8).value
    sum_i22 = ws.cell(22, 9).value
    sum_preserved = (
        isinstance(sum_h22, str) and sum_h22.startswith('=SUM') and
        isinstance(sum_i22, str) and sum_i22.startswith('=SUM')
    )

    # 4. Verify merged ranges preserved (V04 has 48 canonical)
    merged_count = len(ws.merged_cells.ranges)

    # 5. Save
    wb.save(str(output))

    # 6. Post-render validation
    validation = _validate_v04_post_render(str(output), pm37_data)

    return {
        "output_path": str(output),
        "template_used": str(template),
        "schema_version_consumed": schema_version,
        "format_canon_consumed": format_canon,
        "rap_rows_written": rap_rows_written,
        "data_cells_written": data_cells_written,
        "metadata_cells_written": metadata_written,
        "sum_formulas_preserved": sum_preserved,
        "merged_ranges_preserved": merged_count,
        "validation_post_render": validation,
    }


# ═══════════════════════════════════════════════════════════════════════════
# V04 Helpers privados
# ═══════════════════════════════════════════════════════════════════════════

def _populate_v04_metadata(ws, pm37_data):
    """Populate E6-E11 metadata cells V04. Returns count cells written."""
    md = pm37_data.get("programa_metadata", {})
    integrantes = md.get("integrantes_diseno", []) or []

    values = {
        "fecha_elaboracion": md.get("fecha_elaboracion", pm37_data.get("generated_date", "")),
        "denominacion": md.get("denominacion", ""),
        "modalidad": md.get("modalidad", ""),
        "codigo_version": md.get("codigo_version", ""),
        "instructor_1": integrantes[0] if len(integrantes) > 0 else "",
        "instructor_2": integrantes[1] if len(integrantes) > 1 else "",
    }

    written = 0
    for key, cell_coord in V04_METADATA_CELLS.items():
        v = values.get(key, "")
        if v:
            ws[cell_coord] = str(v)
            written += 1
    return written


def _populate_v04_data_rows(ws, rows):
    """Populate data rows R15/R18/R20/R21 from rows[] · one per RAP.

    REGLA 10 single-guía absorpción: RA1 row contains full data · RA2..RAN
    rows only have rap_titulo populated (cols 3-14 shared with RA1 · NOT re-written).

    Returns:
        (rap_rows_written, total_data_cells_written)
    """
    rap_rows_written = 0
    total_cells = 0

    for row_data in rows:
        rap_id = row_data.get("rap_id", "")
        target_row = V04_RAP_ROW.get(rap_id)
        if not target_row:
            continue  # Unknown RAP · skip

        rap_rows_written += 1
        is_ra1 = (rap_id == "RA1")

        for col_field, col_idx in V04_DATA_COLUMN_INDICES.items():
            val = row_data.get(col_field)
            if val is None or val == "":
                continue

            # Skip horas if 0 (let SUM formulas resolve)
            if col_field in ("col_7_horas_trabajo_directo", "col_8_horas_trabajo_independiente") and val == 0:
                continue

            # REGLA 10: For RA2/RA3/RA4 only populate col 2 (rap_titulo)
            # Other cols stay empty (V04 convention · single-guía absorpción)
            if not is_ra1 and col_idx != 2:
                continue

            # Format value · preserve numeric type for horas cells (Excel SUM formulas need numeric)
            if col_field in ("col_7_horas_trabajo_directo", "col_8_horas_trabajo_independiente"):
                try:
                    formatted = int(val)
                except (ValueError, TypeError):
                    formatted = val
            elif isinstance(val, list):
                formatted = "\n".join(str(v) for v in val if v)
            elif isinstance(val, dict):
                formatted = "\n".join(f"{k}: {v}" for k, v in val.items() if v)
            else:
                formatted = str(val)

            ws.cell(target_row, col_idx).value = formatted
            total_cells += 1

    return rap_rows_written, total_cells


def _validate_v04_post_render(output_path, pm37_data):
    """Re-read V04 output xlsx · verify critical assertions.

    Returns dict with check results · status PASS/FAIL/WARN per check.
    """
    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws = wb.active

    checks = {}

    # Check 1: metadata E6-E11 populated (where data available)
    md = pm37_data.get("programa_metadata", {})
    metadata_expected = sum([
        bool(md.get("fecha_elaboracion") or pm37_data.get("generated_date")),
        bool(md.get("denominacion")),
        bool(md.get("modalidad")),
        bool(md.get("codigo_version")),
        len(md.get("integrantes_diseno", [])) >= 1,
        len(md.get("integrantes_diseno", [])) >= 2,
    ])
    metadata_actual = sum([
        bool(ws[V04_METADATA_CELLS[k]].value) for k in V04_METADATA_CELLS.keys()
    ])
    checks["metadata_cells_populated"] = f"{metadata_actual}/6 (expected ≥{metadata_expected})"
    checks["metadata_pass"] = metadata_actual >= metadata_expected

    # Check 2: RAP rows populated (RA1 col_1 + col_2 minimum)
    ra1_row = V04_RAP_ROW["RA1"]
    ra1_competencia = ws.cell(ra1_row, 1).value
    ra1_rap = ws.cell(ra1_row, 2).value
    checks["ra1_competencia_present"] = bool(ra1_competencia)
    checks["ra1_rap_titulo_present"] = bool(ra1_rap)

    # Check 3: Multi-RAP titles (if rows > 1)
    rows = pm37_data.get("rows", [])
    rap_titles_found = 0
    for r in rows:
        rid = r.get("rap_id", "")
        target = V04_RAP_ROW.get(rid)
        if target and ws.cell(target, 2).value:
            rap_titles_found += 1
    checks["rap_titles_populated"] = f"{rap_titles_found}/{len(rows)}"
    checks["rap_titles_pass"] = rap_titles_found == len(rows)

    # Check 4: Horas RA1 populated
    horas_directo = ws.cell(ra1_row, 8).value
    horas_indep = ws.cell(ra1_row, 9).value
    checks["horas_ra1_directo"] = horas_directo
    checks["horas_ra1_indep"] = horas_indep
    totals = pm37_data.get("totals", {})
    expected_directo = totals.get("horas_directas_total")
    expected_indep = totals.get("horas_independientes_total")
    checks["horas_match_totals"] = (
        horas_directo == expected_directo and horas_indep == expected_indep
    )

    # Check 5: SUM formulas R22 preserved
    sum_h22 = ws.cell(22, 8).value
    sum_i22 = ws.cell(22, 9).value
    checks["sum_formula_h22_preserved"] = isinstance(sum_h22, str) and sum_h22.startswith('=SUM')
    checks["sum_formula_i22_preserved"] = isinstance(sum_i22, str) and sum_i22.startswith('=SUM')

    # Check 6: Merged ranges (V04 canon = 48)
    checks["merged_ranges_count"] = len(ws.merged_cells.ranges)
    checks["merged_ranges_pass"] = checks["merged_ranges_count"] == 48

    # Veredicto
    pass_checks = [
        checks["metadata_pass"],
        checks["ra1_competencia_present"],
        checks["ra1_rap_titulo_present"],
        checks["rap_titles_pass"],
        checks["horas_match_totals"],
        checks["sum_formula_h22_preserved"],
        checks["sum_formula_i22_preserved"],
        checks["merged_ranges_pass"],
    ]
    checks["veredicto"] = "PASS" if all(pass_checks) else "FAIL"
    checks["pass_count"] = sum(pass_checks)
    checks["total_checks"] = len(pass_checks)

    return checks

# ═══════════════════════════════════════════════════════════════════════════
# CLI · self-test
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 4:
        print("Uso: python xlsx_renderer.py <pm37_json_path> <pm0_context_json_path> <output_xlsx_path> [template_path] [repo_root]")
        print()
        print("Ejemplo:")
        print("  python xlsx_renderer.py runs/IMARPOR-CC-2026-04-27/pm-3-7.json \\")
        print("                            runs/IMARPOR-CC-2026-04-27/pm-0-context.json \\")
        print("                            runs/IMARPOR-CC-2026-04-27/pm-3-7-gfpi-f134-matrix.xlsx")
        sys.exit(1)

    pm37_path = sys.argv[1]
    pm0_path = sys.argv[2]
    output = sys.argv[3]
    template = sys.argv[4] if len(sys.argv) > 4 else CANON_TEMPLATE_PATH
    repo_root = sys.argv[5] if len(sys.argv) > 5 else None

    pm37_data = json.loads(Path(pm37_path).read_text(encoding="utf-8"))
    pm0_context = json.loads(Path(pm0_path).read_text(encoding="utf-8"))

    print(f"=== xlsx_renderer · v1.0 ===")
    print(f"  pm37 source: {pm37_path}")
    print(f"  pm0 source:  {pm0_path}")
    print(f"  template:    {template}")
    print(f"  output:      {output}")
    print()

    result = render_gfpi_f134_matrix(
        pm37_data=pm37_data,
        pm0_context=pm0_context,
        output_path=output,
        template_path=template,
        repo_root=repo_root,
    )

    print(f"=== Render complete ===")
    print(json.dumps(result, indent=2, ensure_ascii=False))
