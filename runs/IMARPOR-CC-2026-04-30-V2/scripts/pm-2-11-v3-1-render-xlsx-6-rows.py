#!/usr/bin/env python3
"""
PM-2.11 v3.1 · Render xlsx GFPI-F-134 V04 con 6 filas tripartitas
==================================================================

Toma el formato V04 referencia Sergio (GFPI-F-134-V04-REFERENCIA-formato-Sergio.xlsx)
y popula 6 filas (rows 15-20) con las 6 sub-filas tripartitas de pm-2-11.json v3.1.

NO una sola fila monolítica · honra patrón canon "agrupado por RA".
"""
import json
import shutil
import openpyxl
from openpyxl.styles import Alignment
from copy import copy
from pathlib import Path

RUN_DIR = Path('/sessions/beautiful-eager-faraday/mnt/fpi-sena-factory/runs/IMARPOR-CC-2026-04-30-V2')
SRC_TEMPLATE = RUN_DIR / 'GFPI-F-134-V04-REFERENCIA-formato-Sergio.xlsx'
OUTPUT = RUN_DIR / 'pm-2-11-GFPI-F-134-V04.xlsx'
BACKUP = RUN_DIR / 'pm-2-11-GFPI-F-134-V04.pre-v3-1-1-row.xlsx'

# Backup del v3.0 (1 fila) si no existe
if OUTPUT.exists() and not BACKUP.exists():
    shutil.copy(OUTPUT, BACKUP)
    print(f'BACKUP creado: {BACKUP.name}')

# Copy template fresco para garantizar formato V04 limpio
shutil.copy(SRC_TEMPLATE, OUTPUT)

# Load pm-2-11 v3.1
pm211 = json.load(open(RUN_DIR / 'pm-2-11.json'))
rows = pm211['gfpi_f134_v04_rows']
assert len(rows) == 6, f'Expected 6 rows, got {len(rows)}'

# === Open xlsx & populate ===
wb = openpyxl.load_workbook(OUTPUT)
ws = wb['PLANEACIÓN POR RAPS']

# === UNMERGE data rows 15-21 (formato V04 referencia tiene celdas merged a través de 7 sub-rows) ===
# Esto es necesario porque queremos 6 filas REALES con datos diferentes por bloque tripartito,
# no 1 fila bloque-competencia con sub-rows segmentadas.
ranges_to_unmerge = []
for mr in list(ws.merged_cells.ranges):
    if mr.min_row >= 15 and mr.max_row <= 21:
        ranges_to_unmerge.append(str(mr))
for r in ranges_to_unmerge:
    ws.unmerge_cells(r)
print(f'Unmerged {len(ranges_to_unmerge)} data-row ranges (rows 15-21) para honrar 6 filas tripartitas')

# Helper formatting
def set_cell(row_num, col_num, value, wrap=True, valign='top'):
    """Set cell value con wrap y vertical-top. Skip silently si es MergedCell."""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row_num, column=col_num)
    if isinstance(cell, MergedCell):
        # MergedCell: find the top-left of the merge range and write there
        for merge_range in ws.merged_cells.ranges:
            if cell.coordinate in merge_range:
                top_left = ws.cell(row=merge_range.min_row, column=merge_range.min_col)
                if not isinstance(top_left, MergedCell):
                    top_left.value = value
                    top_left.alignment = Alignment(wrap_text=wrap, vertical=valign, horizontal='left')
                return
        return  # Could not find writable cell
    cell.value = value
    cell.alignment = Alignment(wrap_text=wrap, vertical=valign, horizontal='left')

def fmt_list(items, prefix='• '):
    """Convert list → newline-separated string with prefix."""
    if not items:
        return ''
    return '\n'.join(f'{prefix}{str(i)}' for i in items)

def fmt_criterios(c6):
    """Format C6 criterios dict → text."""
    out = []
    if c6.get('sofia_assigned'):
        out.append('— SOFÍA SENA (matriz v1.3) —')
        out.extend(f'  • {c}' for c in c6['sofia_assigned'])
    if c6.get('canon_sistema_assigned'):
        out.append('')
        out.append('— CANON SISTEMA C01-C08 —')
        canon = c6['canon_sistema_assigned']
        if isinstance(canon, list):
            out.extend(f'  • {str(c)}' for c in canon)
        else:
            out.append(f'  • {canon}')
    if c6.get('_rationale'):
        out.append('')
        out.append(f"[Rationale]: {c6['_rationale']}")
    return '\n'.join(out)

def fmt_rap(c2):
    """Format C2 rap dict → text. Skip empty fields."""
    parts = [c2['rap_id']]
    cod = c2.get('codigo', '').strip()
    enu = c2.get('enunciado', '').strip()
    if cod and cod != 'N/A':
        parts.append(cod)
    if enu:
        parts.append(enu)
    return '\n'.join(parts)

def fmt_material(m):
    """Extract text from material entry (could be dict, string, etc.)."""
    if isinstance(m, dict):
        # Prefer descripcion · then nombre · then any text-ish field
        for k in ['descripcion', 'description', 'nombre', 'name', 'item']:
            if m.get(k):
                return str(m[k])
        # Fallback: concatenate all string values
        return ' · '.join(str(v) for v in m.values() if isinstance(v, (str, int, float)))[:200]
    return str(m)

# === Row mapping: pm-2-11 row N → xlsx row 15+N ===
START_ROW = 15

# Need to insert 5 extra rows (template has only 1 data row at 15)
# Use insert_rows or copy row 15 styles to 16-20
for i in range(1, 6):  # rows 16-20
    new_row = START_ROW + i
    # Copy height
    ws.row_dimensions[new_row].height = 600

# Set heights for all 6 data rows
for i in range(6):
    ws.row_dimensions[START_ROW + i].height = 600

# === Populate 6 rows ===
for idx, row_data in enumerate(rows):
    xlsx_row = START_ROW + idx

    # C1 Competencia
    set_cell(xlsx_row, 1, row_data['C1_competencia'])

    # C2 RAP
    set_cell(xlsx_row, 2, fmt_rap(row_data['C2_rap']))

    # C3 (vacío en V04 referencia)
    set_cell(xlsx_row, 3, '')

    # C4 Saberes Conceptos
    set_cell(xlsx_row, 4, fmt_list(row_data['C4_saberes_conceptos']))

    # C5 Saberes Proceso
    set_cell(xlsx_row, 5, fmt_list(row_data['C5_saberes_proceso']))

    # C6 Criterios Evaluación
    set_cell(xlsx_row, 6, fmt_criterios(row_data['C6_criterios_evaluacion']))

    # C7 Actividades Aprendizaje (formato canon Sergio: items multi-línea separados por línea en blanco)
    set_cell(xlsx_row, 7, '\n\n'.join(row_data['C7_actividades_aprendizaje']))

    # C8 Horas Trabajo Directo
    set_cell(xlsx_row, 8, str(row_data['C8_horas_directo']))

    # C9 Horas Trabajo Independiente
    set_cell(xlsx_row, 9, str(row_data['C9_horas_independiente']))

    # C10 Descripción Evidencia (formato canon Sergio: 3-4 líneas por evidencia · separación en blanco)
    set_cell(xlsx_row, 10, '\n\n'.join(row_data['C10_evidencias']))

    # C11 Estrategias Didácticas (formato canon Sergio: por actividad · 3 líneas · separación en blanco)
    set_cell(xlsx_row, 11, '\n\n'.join(row_data['C11_estrategias_didacticas']))

    # C12 Ambientes (texto único transversal)
    set_cell(xlsx_row, 12, str(row_data['C12_ambiente'])[:1000])

    # C13 Materiales
    materiales = row_data['C13_materiales_formacion']
    if isinstance(materiales, list):
        materiales_text = [fmt_material(m) for m in materiales[:30]]  # cap a 30
        set_cell(xlsx_row, 13, fmt_list(materiales_text))
    else:
        set_cell(xlsx_row, 13, str(materiales))

    # C14 Instructores
    set_cell(xlsx_row, 14, str(row_data['C14_instructores_responsables'])[:600])

    # C15 Observaciones
    set_cell(xlsx_row, 15, str(row_data['C15_observaciones']))

    print(f'  Row xlsx {xlsx_row} · {row_data["_bloque_id"]} {row_data["_tipo_bloque"]:>15} · '
          f'RAP={row_data["C2_rap"]["rap_id"]:>22} · '
          f'C7={row_data["_C7_total"]} act · C10={row_data["_C10_total"]} ev')

# Save
wb.save(OUTPUT)
print()
print(f'✅ XLSX V04 v3.1 generado · {OUTPUT.name}')
print(f'   Tamaño: {OUTPUT.stat().st_size // 1024} KB')
print(f'   Filas pobladas: 15-20 (6 filas tripartitas) · NO solo row 15')
