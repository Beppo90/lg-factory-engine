#!/usr/bin/env python3
"""
Wave C · Re-render pm-2-11.json + xlsx con criterios_por_actividad heredados
==============================================================================

Modifica pm-2-11.json v3.1 (v3.3 schema):
- ADD campo top-level `criterios_por_actividad` dict {numero_acumulado: [criterios]}
- ADD a cada row[] gfpi_f134_v04_rows el sub-campo `_criterios_por_actividad_bloque` (subset relevante)
- Enriquece C6_criterios_evaluacion con sección NEW "criterios_per_activity_card" (lista por activity_card)

Re-render xlsx pm-2-11-GFPI-F-134-V04.xlsx con C6 enriquecido (criterios SOFÍA + canon + per-activity).

Backups: pm-2-11.json.pre-wave-c · pm-2-11-GFPI-F-134-V04.pre-wave-c.xlsx
"""
import json
import shutil
from pathlib import Path

RUN_DIR = Path('/sessions/beautiful-eager-faraday/mnt/fpi-sena-factory/runs/IMARPOR-CC-2026-04-30-V2')

# === LOAD INPUTS ===
print('=== Wave C · Re-render pm-2-11 con criterios_por_actividad ===\n')

pm211_path = RUN_DIR / 'pm-2-11.json'
backup_pm211 = RUN_DIR / 'pm-2-11.json.pre-wave-c'

if not backup_pm211.exists():
    shutil.copy(pm211_path, backup_pm211)
    print(f'BACKUP: {backup_pm211.name}')

pm211 = json.load(open(pm211_path))

# Load 30 cards from regenerated PMs
PM_FILES = ['pm-2-1.json','pm-2-2.json','pm-2-3.json','pm-2-4.json','pm-2-5.json',
            'pm-2-6.json','pm-2-8.json','pm-2-9.json','pm-2-10.json','pm-3-5.json','pm-4-2.json']
all_cards = []
for f in PM_FILES:
    j = json.load(open(RUN_DIR / f))
    cards = j.get('activity_cards', j.get('actividades', []))
    if not cards and 'activity_card' in j:
        cards = [j['activity_card']]
    if not cards:
        for k, v in j.items():
            if isinstance(v, list) and v and isinstance(v[0], dict) and ('pm_id' in v[0] or 'session' in v[0] or 'tipo_bloque' in v[0]):
                cards = v
                break
    for c in cards:
        c['_source_pm'] = f
        all_cards.append(c)

print(f'Cards loaded: {len(all_cards)}/30')
cards_with_criterios = [c for c in all_cards if c.get('criterios_evaluacion')]
print(f'Cards con criterios_evaluacion[]: {len(cards_with_criterios)}/30')

# === BUILD criterios_por_actividad dict ===
criterios_por_actividad = {}
for c in all_cards:
    num = str(c.get('numero_actividad', '?'))
    criterios = c.get('criterios_evaluacion', [])
    bloque = c.get('bloque_id_referencia', '?')
    rap = c.get('rap_target') or '-'
    dim = c.get('dimension', '?')
    enun = c.get('enunciado', '')
    alignment = c.get('_alineamiento_criterio_sofia', '-')
    if criterios:
        criterios_por_actividad[num] = {
            'numero_actividad': num,
            'session': str(c.get('session', '?')).replace('SS','').replace('S',''),
            'bloque_id_referencia': bloque,
            'rap_target': rap,
            'dimension': dim,
            'enunciado_short': enun[:120] + ('...' if len(enun) > 120 else ''),
            'criterios': criterios,
            '_alineamiento_sofia': alignment,
        }

print(f'criterios_por_actividad keys: {len(criterios_por_actividad)}')

# === ADD a top-level pm-2-11 ===
pm211['criterios_por_actividad'] = criterios_por_actividad
pm211['_v3_3_added'] = {
    'fecha': '2026-05-02',
    'cascade_source': 'AC v3.1 .criterios_evaluacion[] heredado de las 30 cards regeneradas Wave B',
    'consumo_downstream': 'PM-3.6 v3.3 Sección 4 tabla Col 5 Criterios de Evaluación',
    'method': 'mecanico_deterministic_v3.1 (verbo enunciado → verbo SOFÍA por dimension)',
}

# === ADD a cada row[] gfpi_f134_v04_rows el sub-campo _criterios_por_actividad_bloque ===
rows = pm211.get('gfpi_f134_v04_rows', [])
for row in rows:
    bid = row.get('_bloque_id', '?')
    # Cards de este bloque
    bloque_criterios = {num: data for num, data in criterios_por_actividad.items()
                       if data['bloque_id_referencia'] == bid}
    row['_criterios_por_actividad_bloque'] = bloque_criterios

    # Enriquecer C6_criterios_evaluacion con sub-key NEW
    c6 = row.get('C6_criterios_evaluacion', {})
    if isinstance(c6, dict):
        c6['criterios_per_activity_card'] = [
            f"#{num} ({data['dimension']}): {' | '.join(data['criterios'])}"
            for num, data in sorted(bloque_criterios.items(), key=lambda x: int(x[0]))
        ]
        c6['_v3_3_per_activity_count'] = len(bloque_criterios)

# === Update pm_version ===
pm211['pm_version'] = '3.3'
pm211['_paradigm'] = pm211.get('_paradigm', '') + ' | v3.3: criterios_por_actividad heredados de AC v3.1'

# === SAVE pm-2-11.json ===
with open(pm211_path, 'w', encoding='utf-8') as f:
    json.dump(pm211, f, ensure_ascii=False, indent=2)
print(f'\n✅ pm-2-11.json v3.3 saved · {pm211_path.stat().st_size // 1024} KB')

# Stats por bloque
print(f'\nDistribución criterios_por_actividad por bloque:')
for row in rows:
    bid = row.get('_bloque_id', '?')
    tipo = row.get('_tipo_bloque', '?')
    rap = row.get('C2_rap', {}).get('rap_id', '-')
    n_act = len(row.get('_criterios_por_actividad_bloque', {}))
    print(f'  {bid} {tipo:>15} RAP={rap:>22} · {n_act} activities con criterios')

# === RE-RENDER xlsx ===
print(f'\n=== Re-render xlsx con C6 enriquecido ===')
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell

xlsx_path = RUN_DIR / 'pm-2-11-GFPI-F-134-V04.xlsx'
backup_xlsx = RUN_DIR / 'pm-2-11-GFPI-F-134-V04.pre-wave-c.xlsx'

if not backup_xlsx.exists():
    shutil.copy(xlsx_path, backup_xlsx)
    print(f'BACKUP xlsx: {backup_xlsx.name}')

# Load template fresh y re-renderizar
SRC_TEMPLATE = RUN_DIR / 'GFPI-F-134-V04-REFERENCIA-formato-Sergio.xlsx'
shutil.copy(SRC_TEMPLATE, xlsx_path)

wb = openpyxl.load_workbook(xlsx_path)
ws = wb['PLANEACIÓN POR RAPS']

# Unmerge data rows 15-21
ranges_to_unmerge = []
for mr in list(ws.merged_cells.ranges):
    if mr.min_row >= 15 and mr.max_row <= 21:
        ranges_to_unmerge.append(str(mr))
for r in ranges_to_unmerge:
    ws.unmerge_cells(r)

def set_cell(row_num, col_num, value, wrap=True, valign='top'):
    cell = ws.cell(row=row_num, column=col_num)
    if isinstance(cell, MergedCell):
        for merge_range in ws.merged_cells.ranges:
            if cell.coordinate in merge_range:
                top_left = ws.cell(row=merge_range.min_row, column=merge_range.min_col)
                if not isinstance(top_left, MergedCell):
                    top_left.value = value
                    top_left.alignment = Alignment(wrap_text=wrap, vertical=valign, horizontal='left')
                return
        return
    cell.value = value
    cell.alignment = Alignment(wrap_text=wrap, vertical=valign, horizontal='left')

def fmt_list(items, prefix='• '):
    if not items:
        return ''
    return '\n'.join(f'{prefix}{str(i)}' for i in items)

def fmt_rap(c2):
    parts = [c2['rap_id']]
    cod = c2.get('codigo', '').strip()
    enu = c2.get('enunciado', '').strip()
    if cod and cod != 'N/A':
        parts.append(cod)
    if enu:
        parts.append(enu)
    return '\n'.join(parts)

def fmt_criterios_v3_3(c6):
    """Format C6 con SOFÍA + canon + per-activity criterios."""
    out = []
    if c6.get('sofia_assigned'):
        out.append('— SOFÍA SENA (matriz v1.3) —')
        for c in c6['sofia_assigned']:
            out.append(f'  • {c}')
    if c6.get('canon_sistema_assigned'):
        out.append('')
        out.append('— CANON SISTEMA C01-C08 —')
        canon = c6['canon_sistema_assigned']
        if isinstance(canon, list):
            for c in canon:
                out.append(f'  • {str(c)}')
        else:
            out.append(f'  • {canon}')
    if c6.get('criterios_per_activity_card'):
        out.append('')
        out.append(f'— PER-ACTIVITY (heredados AC v3.1 · {c6.get("_v3_3_per_activity_count", 0)} cards) —')
        for c in c6['criterios_per_activity_card']:
            out.append(f'  • {c}')
    if c6.get('_rationale'):
        out.append('')
        out.append(f"[Rationale]: {c6['_rationale']}")
    return '\n'.join(out)

def fmt_material(m):
    if isinstance(m, dict):
        for k in ['descripcion', 'description', 'nombre', 'name', 'item']:
            if m.get(k):
                return str(m[k])
        return ' · '.join(str(v) for v in m.values() if isinstance(v, (str, int, float)))[:200]
    return str(m)

# Helper functions Wave 5.D canon Sergio C7+C10+C11
def fmt_actividad(c):
    num = c.get('numero_actividad', '?')
    dim = c.get('dimension', 'cognitiva').lower()
    enun = c.get('enunciado') or c.get('titulo') or '?'
    return f"{num}. Actividad {dim}:\n{enun}"

def fmt_evidencia_canon(c):
    ev = c.get('evidencias')
    if not ev or ev is False or not isinstance(ev, dict):
        return None
    if not ev.get('aplica', False):
        return None
    tipo = (ev.get('tipo') or '').lower()
    nombre = ev.get('nombre', '?')
    tecnica = ev.get('tecnica_evaluacion', '?')
    instr_num = ev.get('instrumento_numero', '?')
    instr_tipo = ev.get('instrumento_tipo', '?')
    codigo = ev.get('codigo_canon', '')
    num_act = c.get('numero_actividad', '?')
    header = f"[Actividad {num_act} · {codigo}]" if codigo else f"[Actividad {num_act}]"
    return (f"{header}\n"
            f"Evidencia de {tipo}: {nombre}\n"
            f"Técnica de evaluación: {tecnica}\n"
            f"Instrumento de evaluación No {instr_num}: {instr_tipo}")

def fmt_estrategia_canon(c):
    num = c.get('numero_actividad', '?')
    estrategias = c.get('estrategias_didacticas_activas', c.get('estrategias', []))
    tecnicas = c.get('tecnicas_didacticas', c.get('tecnicas', []))
    if not isinstance(estrategias, list):
        estrategias = [str(estrategias)] if estrategias else []
    if not isinstance(tecnicas, list):
        tecnicas = [str(tecnicas)] if tecnicas else []
    if not estrategias and not tecnicas:
        return None
    est_str = ' + '.join(estrategias) if estrategias else '(no declarada)'
    tec_str = ' + '.join(tecnicas) if tecnicas else '(no declarada)'
    return (f"Actividad {num}.\n"
            f"Estrategia didáctica: {est_str}\n"
            f"Técnica Didáctica: {tec_str}")

# === Build cards by bloque (re-extract with criterios) ===
cards_por_bloque = {'B0': [], 'B1': [], 'B2': [], 'B3': [], 'B4': [], 'BT': []}
for c in all_cards:
    bid = c.get('bloque_id_referencia', '?')
    if bid in cards_por_bloque:
        cards_por_bloque[bid].append(c)
for bid, cs in cards_por_bloque.items():
    cs.sort(key=lambda c: int(str(c.get('numero_actividad', 0)).replace('SS','').replace('S','')))

# === Populate 6 rows ===
START_ROW = 15
for i in range(6):
    ws.row_dimensions[START_ROW + i].height = 700  # más alto para criterios per-activity

competencia = pm211['gfpi_f134_v04_rows'][0]['C1_competencia']

for idx, row in enumerate(pm211['gfpi_f134_v04_rows']):
    xlsx_row = START_ROW + idx
    bid = row['_bloque_id']
    cards = cards_por_bloque.get(bid, [])

    set_cell(xlsx_row, 1, competencia)
    set_cell(xlsx_row, 2, fmt_rap(row['C2_rap']))
    set_cell(xlsx_row, 3, '')
    set_cell(xlsx_row, 4, fmt_list(row['C4_saberes_conceptos']))
    set_cell(xlsx_row, 5, fmt_list(row['C5_saberes_proceso']))
    set_cell(xlsx_row, 6, fmt_criterios_v3_3(row['C6_criterios_evaluacion']))
    set_cell(xlsx_row, 7, '\n\n'.join(fmt_actividad(c) for c in cards))
    set_cell(xlsx_row, 8, str(row['C8_horas_directo']))
    set_cell(xlsx_row, 9, str(row['C9_horas_independiente']))
    evid_list = [e for e in [fmt_evidencia_canon(c) for c in cards] if e]
    if not evid_list:
        if row['_tipo_bloque'] == 'APERTURA':
            evid_list = ['(N/A — APERTURA NO produce evidencias formales)']
        elif row['_tipo_bloque'] == 'TRANSFERENCIA':
            evid_list = ['(esperado: E-Misión Pre-Departure Banana Reefer Compliance Check)']
        else:
            evid_list = ['(verificar)']
    set_cell(xlsx_row, 10, '\n\n'.join(evid_list))
    estr_list = [s for s in [fmt_estrategia_canon(c) for c in cards] if s]
    set_cell(xlsx_row, 11, '\n\n'.join(estr_list) if estr_list else '(heredar pm-3-2 downstream)')

    # C12-C14 desde row data
    set_cell(xlsx_row, 12, str(row['C12_ambiente'])[:1000])
    materiales = row['C13_materiales_formacion']
    if isinstance(materiales, list):
        set_cell(xlsx_row, 13, fmt_list([fmt_material(m) for m in materiales[:30]]))
    else:
        set_cell(xlsx_row, 13, str(materiales))
    set_cell(xlsx_row, 14, str(row['C14_instructores_responsables'])[:600])
    set_cell(xlsx_row, 15, str(row['C15_observaciones']))

    print(f'  Row xlsx {xlsx_row} · {bid} · {len(cards)} act · C6 sofia+canon+per-activity({len(row["_criterios_por_actividad_bloque"])}) lines')

wb.save(xlsx_path)
print(f'\n✅ XLSX V04 v3.3 saved · {xlsx_path.stat().st_size // 1024} KB')
print(f'\n=== Wave C COMPLETO · listo para Wave D ===')
